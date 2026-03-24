# ============================================================
# Module 2: Create Excel Sample Files
# ============================================================
# This script generates .xlsx workbooks from the course CSV
# datasets. The workbooks include formatted headers, column
# widths, and a few pre-built formulas so students can open
# them and start exploring immediately.
#
# Usage:
#   python create_excel_samples.py
#
# Output:
#   ../data/excel/sales_data.xlsx
#   ../data/excel/employees.xlsx
#   ../data/excel/customers.xlsx
# ============================================================

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "data")
OUTPUT_DIR = os.path.join(DATA_DIR, "excel")


def style_workbook(filepath):
    """Apply professional formatting to an Excel workbook."""
    wb = load_workbook(filepath)
    ws = wb.active

    # Style the header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths (approximate)
    for col_num in range(1, ws.max_column + 1):
        max_length = len(str(ws.cell(row=1, column=col_num).value or ""))
        for row_num in range(2, min(ws.max_row + 1, 50)):  # Sample first 50 rows
            val = ws.cell(row=row_num, column=col_num).value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    wb.save(filepath)


def add_summary_sheet(filepath, df, group_col, value_col):
    """Add a Summary sheet with formulas demonstrating Excel functions."""
    wb = load_workbook(filepath)
    ws = wb.create_sheet("Summary")

    # Title
    ws["A1"] = "Quick Summary"
    ws["A1"].font = Font(bold=True, size=14)

    # Basic stats using Excel formulas
    data_sheet = wb.sheetnames[0]
    col_letter = get_column_letter(df.columns.tolist().index(value_col) + 1)
    last_row = len(df) + 1

    stats = [
        ("Total Rows", f'=COUNTA({data_sheet}!A2:A{last_row})'),
        (f"Sum of {value_col}", f'=SUM({data_sheet}!{col_letter}2:{col_letter}{last_row})'),
        (f"Average {value_col}", f'=AVERAGE({data_sheet}!{col_letter}2:{col_letter}{last_row})'),
        (f"Min {value_col}", f'=MIN({data_sheet}!{col_letter}2:{col_letter}{last_row})'),
        (f"Max {value_col}", f'=MAX({data_sheet}!{col_letter}2:{col_letter}{last_row})'),
    ]

    for i, (label, formula) in enumerate(stats, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = formula

    # Note for students
    ws[f"A{len(stats) + 5}"] = "These formulas reference the data sheet."
    ws[f"A{len(stats) + 6}"] = "Click on any cell in column B to see the formula in the formula bar."
    ws[f"A{len(stats) + 5}"].font = Font(italic=True, color="666666")
    ws[f"A{len(stats) + 6}"].font = Font(italic=True, color="666666")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    wb.save(filepath)


def create_workbook(csv_path, xlsx_path, group_col, value_col):
    """Create a formatted Excel workbook from a CSV file."""
    filename = os.path.basename(xlsx_path)

    df = pd.read_csv(csv_path)
    df.to_excel(xlsx_path, index=False, sheet_name="Data")
    style_workbook(xlsx_path)
    add_summary_sheet(xlsx_path, df, group_col, value_col)

    print(f"  [OK] {filename} ({len(df)} rows, {len(df.columns)} columns + Summary sheet)")


def main():
    print("=" * 60)
    print("  Creating Excel Sample Files")
    print("=" * 60)
    print()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Sales data
    create_workbook(
        os.path.join(DATA_DIR, "sales_data.csv"),
        os.path.join(OUTPUT_DIR, "sales_data.xlsx"),
        group_col="category",
        value_col="revenue",
    )

    # Employees
    create_workbook(
        os.path.join(DATA_DIR, "employees.csv"),
        os.path.join(OUTPUT_DIR, "employees.xlsx"),
        group_col="department",
        value_col="salary",
    )

    # Customers
    create_workbook(
        os.path.join(DATA_DIR, "customers.csv"),
        os.path.join(OUTPUT_DIR, "customers.xlsx"),
        group_col="state",
        value_col="total_spent",
    )

    print()
    print("=" * 60)
    print("  Done! Files saved to data/excel/")
    print("=" * 60)
    print()
    print("Open these files in Excel (or Google Sheets) to start Module 2.")


if __name__ == "__main__":
    main()
